"""
Excel Sync Service — Dynamic Excel Generation entirely from scratch using openpyxl.
"""
import os
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlalchemy.orm import Session

from ..models import Project, Material, Document, Delivery, COAdjustment, VPO, Inventory
from ..config import EXPORT_DIR, LEGACY_EXCEL

# ──────────────────────────────────────────────────────────────────────────────
# Sheet / column layout (exact mapping from generate_requirements.py)
# ──────────────────────────────────────────────────────────────────────────────

SHEET_VPOS = "VPO's"
SHEET_COBIA = "Cobia Cove Appartments"
SHEET_WILLOW = "Willow Way Apts"

COBIA_COLS = {
    "type": "A", "qty": "B", "co_qty": "AP", "po_co_qty": "AQ",
    "thickness": "AR", "x": "AS", "width": "AT", "length": "AU",
    "material_type": "AV", "lf_pcs": "AW", "bf_sf": "AX",
    "cost_mbf": "AY", "total_cost": "AZ", "total_cost_tax": "BA",
    "invoice_num": "BB",
    "delivery_start": "BC",   # First delivery-date column (row 2 has the date)
    "total_delivered": "DV", "delivered_lf": "DW",
    "delivered_bf_sf": "DX", "delivered_cost": "DY", "delivered_cost_tax": "DZ",
    "pct_delivery": "EA", "inv_bundles": "EB", "inv_uom": "EC",
    "pcs_bundle": "ED", "inv_pcs": "EE", "issues": "EF",
    "issues_lf": "EG", "issues_bf_sf": "EH", "pct_issued": "EI",
    "issues_cost": "EJ", "issues_cost_tax": "EK",
    "variance_code": "EL", "reason": "EM",
}

WILLOW_COLS = {
    "type": "A", "qty": "B", "co_qty": "Q", "po_co_qty": "R",
    "thickness": "S", "x": "T", "width": "U", "length": "V",
    "material_type": "W", "lf_pcs": "X", "bf_sf": "Y",
    "cost_mbf": "Z", "total_cost": "AA", "total_cost_tax": "AB",
    "invoice_num": "AC",
    "delivery_start": "AD",
    "total_delivered": "AV", "delivered_lf": "AW",
    "delivered_bf_sf": "AX", "delivered_cost": "AY", "delivered_cost_tax": "AZ",
    "pct_delivery": "BA", "inv_bundles": "BB", "inv_uom": "BC",
    "pcs_bundle": "BD", "inv_pcs": "BE", "issues": "BF",
    "issues_lf": "BG", "issues_bf_sf": "BH", "pct_issued": "BI",
    "issues_cost": "BJ", "issues_cost_tax": "BK",
    "variance_code": "BL", "reason": "BM",
}


# ──────────────────────────────────────────────────────────────────────────────
# Helpers & Styling
# ──────────────────────────────────────────────────────────────────────────────

def col_to_num(col_letter: str) -> int:
    result = 0
    for c in col_letter.upper():
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result


def natural_sort_key(s):
    if not s:
        return []
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(s))]


# Font name
font_name = "Segoe UI"

# Styles
regular_font = Font(name=font_name, size=10)
bold_font = Font(name=font_name, size=10, bold=True)
header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
sub_header_font = Font(name=font_name, size=10, bold=True, color="1F497D")

# Fills
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

# Borders
thin_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")


def setup_section_headers(ws, is_cobia: bool):
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    
    if is_cobia:
        sections = [
            ("Materials", "A", "B"),
            ("Change Orders", "C", "AO"),
            ("Info", "AP", "AQ"),
            ("Specs", "AR", "BA"),
            ("Invoices", "BB", "BB"),
            ("Deliveries", "BC", "DU"),
            ("Delivery Calculations", "DV", "DZ"),
            ("Delivery %", "EA", "EA"),
            ("Inventory", "EB", "ED"),
            ("Inventory Calculations", "EE", "EK"),
            ("Comments", "EL", "EM")
        ]
    else:
        sections = [
            ("Materials", "A", "B"),
            ("Change Orders", "C", "P"),
            ("Info", "Q", "R"),
            ("Specs", "S", "AB"),
            ("Invoices", "AC", "AC"),
            ("Deliveries", "AD", "AU"),
            ("Delivery Calculations", "AV", "AZ"),
            ("Delivery %", "BA", "BA"),
            ("Inventory", "BB", "BD"),
            ("Inventory Calculations", "BE", "BK"),
            ("Comments", "BL", "BM")
        ]
        
    for label, start_col, end_col in sections:
        start_idx = col_to_num(start_col)
        end_idx = col_to_num(end_col)
        if start_idx == end_idx:
            cell = ws.cell(row=1, column=start_idx, value=label)
        else:
            ws.merge_cells(start_row=1, start_column=start_idx, end_row=1, end_column=end_idx)
            cell = ws.cell(row=1, column=start_idx, value=label)
            
        for r in range(1, 2):
            for c in range(start_idx, end_idx + 1):
                cur_cell = ws.cell(row=r, column=c)
                cur_cell.fill = header_fill
                cur_cell.font = header_font
                cur_cell.alignment = align_center
                cur_cell.border = thin_border


def get_row2_headers(is_cobia: bool, co_labels: List[str]) -> Dict[int, str]:
    headers = {}
    cols = COBIA_COLS if is_cobia else WILLOW_COLS
    
    headers[col_to_num(cols["type"])] = "Type"
    headers[col_to_num(cols["qty"])] = "PO Qty"
    
    co_start = 3
    co_end = 41 if is_cobia else 16
    for c in range(co_start, co_end + 1):
        idx = c - co_start
        if idx < len(co_labels):
            headers[c] = co_labels[idx]
        else:
            headers[c] = f"CO{idx + 1}"
            
    headers[col_to_num(cols["co_qty"])] = "CO Qty"
    headers[col_to_num(cols["po_co_qty"])] = "PO/CO Qty"
    headers[col_to_num(cols["thickness"])] = "Thickness"
    headers[col_to_num(cols["x"])] = "x"
    headers[col_to_num(cols["width"])] = "Width"
    headers[col_to_num(cols["length"])] = "Length"
    headers[col_to_num(cols["material_type"])] = "Material Type"
    headers[col_to_num(cols["lf_pcs"])] = "L/F"
    headers[col_to_num(cols["bf_sf"])] = "B/F - S/F"
    headers[col_to_num(cols["cost_mbf"])] = "Cost/MBF"
    headers[col_to_num(cols["total_cost"])] = "Total Cost"
    headers[col_to_num(cols["total_cost_tax"])] = "Total Cost + Tax"
    headers[col_to_num(cols["invoice_num"])] = "Invoice Num"
    
    headers[col_to_num(cols["total_delivered"])] = "Total Delivered"
    headers[col_to_num(cols["delivered_lf"])] = "Delivered L/F"
    headers[col_to_num(cols["delivered_bf_sf"])] = "Delivered B/F"
    headers[col_to_num(cols["delivered_cost"])] = "Delivered Cost"
    headers[col_to_num(cols["delivered_cost_tax"])] = "Delivered Cost + Tax"
    headers[col_to_num(cols["pct_delivery"])] = "% Delivery"
    
    headers[col_to_num(cols["inv_bundles"])] = "Inv Bundles"
    headers[col_to_num(cols["inv_uom"])] = "Inv UOM"
    headers[col_to_num(cols["pcs_bundle"])] = "PCS/Bundle"
    headers[col_to_num(cols["inv_pcs"])] = "Inv PCS"
    headers[col_to_num(cols["issues"])] = "Issues"
    headers[col_to_num(cols["issues_lf"])] = "Issues L/F"
    headers[col_to_num(cols["issues_bf_sf"])] = "Issues B/F"
    headers[col_to_num(cols["pct_issued"])] = "% Issued"
    headers[col_to_num(cols["issues_cost"])] = "Issues Cost"
    headers[col_to_num(cols["issues_cost_tax"])] = "Issues Cost + Tax"
    headers[col_to_num(cols["variance_code"])] = "Variance Code"
    headers[col_to_num(cols["reason"])] = "Reason"
    
    return headers


def format_data_row(ws, row: int, cols: dict, is_cobia: bool):
    formats = {
        "qty": "#,##0",
        "co_qty": "#,##0",
        "po_co_qty": "#,##0",
        "thickness": "0.0",
        "width": "0.0",
        "length": "0.0",
        "lf_pcs": "#,##0.00",
        "bf_sf": "#,##0.00",
        "cost_mbf": "$#,##0.00",
        "total_cost": "$#,##0.00",
        "total_cost_tax": "$#,##0.00",
        "total_delivered": "#,##0",
        "delivered_lf": "#,##0.00",
        "delivered_bf_sf": "#,##0.00",
        "delivered_cost": "$#,##0.00",
        "delivered_cost_tax": "$#,##0.00",
        "pct_delivery": "0.0%",
        "inv_bundles": "#,##0",
        "pcs_bundle": "#,##0",
        "inv_pcs": "#,##0",
        "issues": "#,##0",
        "issues_lf": "#,##0.00",
        "issues_bf_sf": "#,##0.00",
        "pct_issued": "0.0%",
        "issues_cost": "$#,##0.00",
        "issues_cost_tax": "$#,##0.00"
    }
    
    left_cols = ["type", "material_type", "invoice_num", "inv_uom", "variance_code", "reason"]
    center_cols = ["x"]
    
    for key, col_letter in cols.items():
        col_idx = col_to_num(col_letter)
        cell = ws.cell(row=row, column=col_idx)
        cell.font = regular_font
        cell.border = thin_border
        
        if key in left_cols:
            cell.alignment = align_left
        elif key in center_cols:
            cell.alignment = align_center
        else:
            cell.alignment = align_right
            
        if key in formats:
            cell.number_format = formats[key]
            
    co_end = 41 if is_cobia else 16
    for c in range(3, co_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = align_right
        cell.number_format = "#,##0"
        
    del_start_idx = col_to_num(cols["delivery_start"])
    total_del_idx = col_to_num(cols["total_delivered"])
    for c in range(del_start_idx, total_del_idx):
        cell = ws.cell(row=row, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = align_right
        cell.number_format = "#,##0"


def write_row_formulas_from_scratch(ws, row: int, cols: dict, is_cobia: bool, sorted_dates: list, del_start_idx: int, tax_rate: float):
    r = str(row)
    L = lambda key: cols[key]
    
    type_let = L("type")
    qty_let = L("qty")
    co_qty_let = L("co_qty")
    po_co_qty_let = L("po_co_qty")
    t_let = L("thickness")
    w_let = L("width")
    l_let = L("length")
    cost_let = L("cost_mbf")
    tc_let = L("total_cost")
    tc_tax_let = L("total_cost_tax")
    td_let = L("total_delivered")
    dlf_let = L("delivered_lf")
    dbf_let = L("delivered_bf_sf")
    dc_let = L("delivered_cost")
    dc_tax_let = L("delivered_cost_tax")
    pdel_let = L("pct_delivery")
    ib_let = L("inv_bundles")
    pb_let = L("pcs_bundle")
    ip_let = L("inv_pcs")
    iss_let = L("issues")
    iss_lf_let = L("issues_lf")
    ibf_let = L("issues_bf_sf")
    piss_let = L("pct_issued")
    ic_let = L("issues_cost")
    ic_tax_let = L("issues_cost_tax")
    
    co_end_letter = "AO" if is_cobia else "P"
    
    row_type_val = ws.cell(row=row, column=col_to_num(type_let)).value
    rt = str(row_type_val).strip().lower() if row_type_val else "lumber"
    
    # Formula 1: total_cost
    if rt == "lumber":
        ws.cell(row=row, column=col_to_num(tc_let), value=f"=({qty_let}{r}*{t_let}{r}*{w_let}{r}*{l_let}{r}/12)*{cost_let}{r}/1000")
    elif rt == "panels":
        ws.cell(row=row, column=col_to_num(tc_let), value=f"=({qty_let}{r}*{t_let}{r}*{w_let}{r})*{cost_let}{r}/1000")
    elif rt == "lvl":
        ws.cell(row=row, column=col_to_num(tc_let), value=f"={qty_let}{r}*{l_let}{r}*{cost_let}{r}")
    else:
        ws.cell(row=row, column=col_to_num(tc_let), value=f"={qty_let}{r}*{cost_let}{r}")
        
    # Formula 2: total_cost_tax
    ws.cell(row=row, column=col_to_num(tc_tax_let), value=f"={tc_let}{r}*{tax_rate}")
    
    # Formula 3: co_qty
    ws.cell(row=row, column=col_to_num(co_qty_let), value=f"=SUM(C{r}:{co_end_letter}{r})")
    
    # Formula 4: po_co_qty
    ws.cell(row=row, column=col_to_num(po_co_qty_let), value=f"={qty_let}{r}+{co_qty_let}{r}")
    
    # Formula 5: total_delivered
    if sorted_dates:
        del_start_letter = get_column_letter(del_start_idx)
        del_end_letter = get_column_letter(del_start_idx + len(sorted_dates) - 1)
        ws.cell(row=row, column=col_to_num(td_let), value=f"=SUM({del_start_letter}{r}:{del_end_letter}{r})")
    else:
        del_start_letter = get_column_letter(del_start_idx)
        ws.cell(row=row, column=col_to_num(td_let), value=f"=SUM({del_start_letter}{r}:{del_start_letter}{r})")
        
    # Formula 6: delivered_lf
    ws.cell(row=row, column=col_to_num(dlf_let), value=f"={td_let}{r}*{l_let}{r}")
    
    # Formula 7: delivered_bf_sf
    if rt == "lumber":
        ws.cell(row=row, column=col_to_num(dbf_let), value=f"={td_let}{r}*{t_let}{r}*{w_let}{r}*{l_let}{r}/12")
    elif rt == "panels":
        ws.cell(row=row, column=col_to_num(dbf_let), value=f"={td_let}{r}*{t_let}{r}*{w_let}{r}")
    elif rt == "lvl":
        ws.cell(row=row, column=col_to_num(dbf_let), value=f"={td_let}{r}*{l_let}{r}")
    else:
        ws.cell(row=row, column=col_to_num(dbf_let), value=f"={td_let}{r}")
        
    # Formula 8: delivered_cost
    if rt in ("lumber", "panels"):
        ws.cell(row=row, column=col_to_num(dc_let), value=f"={dbf_let}{r}*{cost_let}{r}/1000")
    elif rt == "lvl":
        ws.cell(row=row, column=col_to_num(dc_let), value=f"={dlf_let}{r}*{cost_let}{r}")
    else:
        ws.cell(row=row, column=col_to_num(dc_let), value=f"={td_let}{r}*{cost_let}{r}")
        
    # Formula 9: delivered_cost_tax
    ws.cell(row=row, column=col_to_num(dc_tax_let), value=f"={dc_let}{r}*{tax_rate}")
    
    # Formula 10: pct_delivery
    ws.cell(row=row, column=col_to_num(pdel_let), value=f"=IFERROR({dc_let}{r}/{tc_let}{r},0)")
    
    # Formula 11: inv_pcs
    ws.cell(row=row, column=col_to_num(ip_let), value=f"=IF({ib_let}{r}<>\"\",{ib_let}{r}*{pb_let}{r},0)")
    
    # Formula 12: issues
    ws.cell(row=row, column=col_to_num(iss_let), value=f"={td_let}{r}-{ip_let}{r}")
    
    # Formula 13: issues_lf
    ws.cell(row=row, column=col_to_num(iss_lf_let), value=f"={iss_let}{r}*{l_let}{r}")
    
    # Formula 14: issues_bf_sf
    if rt == "lumber":
        ws.cell(row=row, column=col_to_num(ibf_let), value=f"=({iss_let}{r}*{t_let}{r}*{w_let}{r}*{l_let}{r})/12")
    elif rt == "panels":
        ws.cell(row=row, column=col_to_num(ibf_let), value=f"={iss_let}{r}*{t_let}{r}*{w_let}{r}")
    else:
        ws.cell(row=row, column=col_to_num(ibf_let), value=f"={iss_let}{r}")
        
    # Formula 15: pct_issued
    ws.cell(row=row, column=col_to_num(piss_let), value=f"=IFERROR({ibf_let}{r}/{td_let}{r},0)")
    
    # Formula 16: issues_cost
    ws.cell(row=row, column=col_to_num(ic_let), value=f"=IFERROR({ibf_let}{r}*{cost_let}{r}/1000,0)")
    
    # Formula 17: issues_cost_tax
    ws.cell(row=row, column=col_to_num(ic_tax_let), value=f"={ic_let}{r}*{tax_rate}")


def write_section_spacers(ws, is_cobia: bool):
    if not is_cobia:
        return
    spacers = [
        (121, "AV", "PANELS SECTION"),
        (155, "AV", "LVL SECTION"),
        (172, "AV", "HARDWARE / MOCKUP SECTION")
    ]
    for r, col_let, text in spacers:
        col_idx = col_to_num(col_let)
        cell = ws.cell(row=r, column=col_idx, value=text)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="1F497D")
        cell.alignment = align_left


def generate_vpos_sheet(ws_vpos, vpos):
    ws_vpos.merge_cells("A1:L1")
    r1_cell = ws_vpos.cell(row=1, column=1, value="Vendor Purchase Orders / Change Orders (VPOs)")
    r1_cell.font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    r1_cell.fill = header_fill
    r1_cell.alignment = align_center
    ws_vpos.row_dimensions[1].height = 25
    
    headers = [
        "Date", "Description", "Qty", "UOM", "Footage", 
        "Price", "Amount", "Tax", "Total", "CO Ref", "CO #", "Remarks"
    ]
    ws_vpos.row_dimensions[2].height = 20
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_vpos.cell(row=2, column=col_idx, value=h)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    formats = {
        3: "#,##0",        # Qty
        5: "#,##0.00",     # Footage
        6: "$#,##0.00",    # Price
        7: "$#,##0.00",    # Amount
        8: "$#,##0.00",    # Tax
        9: "$#,##0.00"     # Total
    }
    
    for row_idx, v in enumerate(vpos, start=3):
        cell_date = ws_vpos.cell(row=row_idx, column=1)
        if v.vpo_date:
            cell_date.value = v.vpo_date
            cell_date.number_format = "yyyy-mm-dd"
        else:
            cell_date.value = ""
            
        ws_vpos.cell(row=row_idx, column=2, value=v.description or "")
        ws_vpos.cell(row=row_idx, column=3, value=v.qty or 0.0)
        ws_vpos.cell(row=row_idx, column=4, value=v.uom or "")
        ws_vpos.cell(row=row_idx, column=5, value=v.footage or 0.0)
        ws_vpos.cell(row=row_idx, column=6, value=v.price or 0.0)
        ws_vpos.cell(row=row_idx, column=7, value=v.amount or 0.0)
        ws_vpos.cell(row=row_idx, column=8, value=v.tax or 0.0)
        ws_vpos.cell(row=row_idx, column=9, value=v.total or 0.0)
        ws_vpos.cell(row=row_idx, column=10, value=v.co_ref or "")
        ws_vpos.cell(row=row_idx, column=11, value=v.co_number or "")
        ws_vpos.cell(row=row_idx, column=12, value=v.remarks or "")
        
        for c in range(1, 13):
            cell = ws_vpos.cell(row=row_idx, column=c)
            cell.font = regular_font
            cell.border = thin_border
            if c in (2, 10, 11, 12):
                cell.alignment = align_left
            elif c == 1:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                
            if c in formats:
                cell.number_format = formats[c]


def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val = cell.value
            if val is not None:
                if isinstance(val, datetime):
                    val_str = val.strftime("%Y-%m-%d")
                else:
                    val_str = str(val)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 10)


# ──────────────────────────────────────────────────────────────────────────────
# Legacy backward compatibility functions
# ──────────────────────────────────────────────────────────────────────────────

def get_delivery_date_columns(ws, start_col: str, row: int = 2) -> Dict[datetime, int]:
    date_cols: Dict[datetime, int] = {}
    start_num = col_to_num(start_col)
    for col in range(start_num, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, datetime):
            date_cols[cell.value] = col
    return date_cols


def find_or_add_date_col(ws, ship_date: datetime, date_cols: Dict[datetime, int], date_row: int = 2) -> int:
    key = datetime(ship_date.year, ship_date.month, ship_date.day)
    if key in date_cols:
        return date_cols[key]
    new_col = max(date_cols.values()) + 1 if date_cols else col_to_num("BC")
    ws.cell(row=date_row, column=new_col, value=key)
    date_cols[key] = new_col
    return new_col


def update_invoice_ref_cell(ws, row: int, col_num: int, invoice_number: str):
    cell = ws.cell(row=row, column=col_num)
    current = str(cell.value) if cell.value else ""
    if invoice_number not in current:
        cell.value = (current + "\n" + invoice_number).strip()


def write_row_formulas(ws, row: int, cols: dict, del_start_letter: str, del_end_letter: str, tax_rate: float = 1.06):
    """Fallback formula writer matching legacy signature."""
    r = str(row)
    td = get_column_letter(col_to_num(cols["total_delivered"]))
    llen = get_column_letter(col_to_num(cols["length"]))
    t = get_column_letter(col_to_num(cols["thickness"]))
    w = get_column_letter(col_to_num(cols["width"]))
    cost = get_column_letter(col_to_num(cols["cost_mbf"]))
    tc = get_column_letter(col_to_num(cols["total_cost"]))
    ib = get_column_letter(col_to_num(cols["inv_bundles"]))
    pb = get_column_letter(col_to_num(cols["pcs_bundle"]))
    ip = get_column_letter(col_to_num(cols["inv_pcs"]))
    iss = get_column_letter(col_to_num(cols["issues"]))
    ibf = get_column_letter(col_to_num(cols["issues_bf_sf"]))
    ic = get_column_letter(col_to_num(cols["issues_cost"]))
    dbf = get_column_letter(col_to_num(cols["delivered_bf_sf"]))
    dc = get_column_letter(col_to_num(cols["delivered_cost"]))

    row_type_val = ws.cell(row=row, column=col_to_num(cols["type"])).value
    rt = str(row_type_val).strip().lower() if row_type_val else ""

    ws.cell(row=row, column=col_to_num(cols["total_delivered"]), value=f"=SUM({del_start_letter}{r}:{del_end_letter}{r})")
    ws.cell(row=row, column=col_to_num(cols["delivered_lf"]), value=f"={td}{r}*{llen}{r}")

    if rt == "lumber":
        ws.cell(row=row, column=col_to_num(cols["delivered_bf_sf"]), value=f"={td}{r}*{t}{r}*{w}{r}*{llen}{r}/12")
    elif rt == "panels":
        ws.cell(row=row, column=col_to_num(cols["delivered_bf_sf"]), value=f"={td}{r}*{t}{r}*{w}{r}")
    elif rt == "lvl":
        ws.cell(row=row, column=col_to_num(cols["delivered_bf_sf"]), value=f"={td}{r}*{llen}{r}")
    else:
        ws.cell(row=row, column=col_to_num(cols["delivered_bf_sf"]), value=f"={td}{r}")

    if rt in ("lumber", "panels"):
        ws.cell(row=row, column=col_to_num(cols["delivered_cost"]), value=f"={dbf}{r}*{cost}{r}/1000")
    elif rt == "lvl":
        ws.cell(row=row, column=col_to_num(cols["delivered_cost"]), value=f"={get_column_letter(col_to_num(cols['delivered_lf']))}{r}*{cost}{r}")
    else:
        ws.cell(row=row, column=col_to_num(cols["delivered_cost"]), value=f"={td}{r}*{cost}{r}")

    ws.cell(row=row, column=col_to_num(cols["delivered_cost_tax"]), value=f"={dc}{r}*{tax_rate}")
    ws.cell(row=row, column=col_to_num(cols["pct_delivery"]), value=f"=IFERROR({dc}{r}/{tc}{r},0)")
    ws.cell(row=row, column=col_to_num(cols["inv_pcs"]), value=f"=IF({ib}{r}<>\"\",{ib}{r}*{pb}{r},0)")
    ws.cell(row=row, column=col_to_num(cols["issues"]), value=f"={td}{r}-{ip}{r}")
    ws.cell(row=row, column=col_to_num(cols["issues_lf"]), value=f"={iss}{r}*{llen}{r}")

    if rt == "lumber":
        ws.cell(row=row, column=col_to_num(cols["issues_bf_sf"]), value=f"=({iss}{r}*{t}{r}*{w}{r}*{llen}{r})/12")
    elif rt == "panels":
        ws.cell(row=row, column=col_to_num(cols["issues_bf_sf"]), value=f"={iss}{r}*{t}{r}*{w}{r}")
    else:
        ws.cell(row=row, column=col_to_num(cols["issues_bf_sf"]), value=f"={iss}{r}")

    ws.cell(row=row, column=col_to_num(cols["pct_issued"]), value=f"=IFERROR({ibf}{r}/{td}{r},0)")
    ws.cell(row=row, column=col_to_num(cols["issues_cost"]), value=f"=IFERROR({ibf}{r}*{cost}{r}/1000,0)")
    ws.cell(row=row, column=col_to_num(cols["issues_cost_tax"]), value=f"={ic}{r}*{tax_rate}")


def _find_excel_row(ws, mat: Material, cols: dict, data_ranges: List[Tuple[int, int]]) -> Optional[int]:
    """
    Find the Excel row that matches this Material DB record by dimensions + type.
    """
    type_col = col_to_num(cols["type"])
    thick_col = col_to_num(cols["thickness"])
    width_col = col_to_num(cols["width"])
    len_col = col_to_num(cols["length"])
    mat_col = col_to_num(cols["material_type"])

    best_row = None
    best_score = 0

    for start_row, end_row in data_ranges:
        for row in range(start_row, end_row + 1):
            row_type = ws.cell(row=row, column=type_col).value
            if not row_type:
                continue

            score = 0
            rt = str(row_type).strip().lower()
            mt = str(mat.type or "").strip().lower()
            if rt == mt:
                score += 10
            elif rt in mt or mt in rt:
                score += 5
            else:
                continue

            def safe_float(v):
                try: return float(str(v).strip())
                except: return None

            row_t = safe_float(ws.cell(row=row, column=thick_col).value)
            row_w = safe_float(ws.cell(row=row, column=width_col).value)
            row_l = safe_float(ws.cell(row=row, column=len_col).value)

            if mat.thickness and row_t == mat.thickness: score += 5
            if mat.width and row_w == mat.width:         score += 5
            if mat.length and row_l == mat.length:       score += 5

            row_mat = str(ws.cell(row=row, column=mat_col).value or "").upper()
            for word in str(mat.material_type or "").upper().split():
                if len(word) > 2 and word in row_mat:
                    score += 2

            if score > best_score:
                best_score = score
                best_row = row

    return best_row if best_score >= 15 else None


# ──────────────────────────────────────────────────────────────────────────────
# Main sync function (from scratch)
# ──────────────────────────────────────────────────────────────────────────────

def sync_excel_for_project(db: Session, project: Project) -> str:
    export_filename = f"{project.name.replace(' ', '_')}_Requirements.xlsx"
    export_path = os.path.join(EXPORT_DIR, export_filename)
    
    name_upper = project.name.upper()
    is_cobia = "COBIA" in name_upper
    
    if is_cobia:
        sheet_name = SHEET_COBIA
        cols = COBIA_COLS
        valid_rows = []
        for r in range(3, 119): valid_rows.append(r)
        for r in range(123, 153): valid_rows.append(r)
        for r in range(157, 171): valid_rows.append(r)
        for r in range(173, 177): valid_rows.append(r)
        co_start_idx = 3
        co_end_idx = 41
    else:
        sheet_name = SHEET_WILLOW
        cols = WILLOW_COLS
        valid_rows = list(range(3, 79))
        co_start_idx = 3
        co_end_idx = 16
        
    # Fetch data
    materials = db.query(Material).filter(Material.project_id == project.id).order_by(Material.id).all()
    material_ids = [m.id for m in materials]
    
    co_adjustments = db.query(COAdjustment).filter(COAdjustment.material_id.in_(material_ids)).all() if material_ids else []
    deliveries = db.query(Delivery).filter(Delivery.material_id.in_(material_ids)).all() if material_ids else []
    vpos = db.query(VPO).filter(VPO.project_id == project.id).all()
    
    # Dynamic headers: Change Orders
    co_numbers = sorted(list(set(adj.co_number for adj in co_adjustments if adj.co_number)), key=natural_sort_key)
    co_col_map = {}
    for i, co_num in enumerate(co_numbers):
        col_num = co_start_idx + i
        if col_num <= co_end_idx:
            co_col_map[co_num] = col_num
            
    # Dynamic headers: Deliveries
    ship_dates = set()
    for d in deliveries:
        if d.ship_date:
            dt = d.ship_date
            if hasattr(dt, 'date'):
                ship_dates.add(datetime(dt.year, dt.month, dt.day))
    sorted_dates = sorted(list(ship_dates))
    
    del_start_idx = col_to_num(cols["delivery_start"])
    total_del_idx = col_to_num(cols["total_delivered"])
    del_end_idx = total_del_idx - 1
    
    date_col_map = {}
    for i, dt in enumerate(sorted_dates):
        col_num = del_start_idx + i
        if col_num <= del_end_idx:
            date_col_map[dt] = col_num
            
    # Create workbook
    wb = openpyxl.Workbook()
    ws_project = wb.active
    ws_project.title = sheet_name
    ws_vpos = wb.create_sheet(title=SHEET_VPOS)
    
    # 1. Setup project sheet section headers
    setup_section_headers(ws_project, is_cobia)
    
    # 2. Setup project sheet row 2 headers
    row2_headers = get_row2_headers(is_cobia, co_numbers)
    for c, label in row2_headers.items():
        ws_project.cell(row=2, column=c, value=label)
        
    for dt, col_num in date_col_map.items():
        cell = ws_project.cell(row=2, column=col_num, value=dt)
        cell.number_format = 'yyyy-mm-dd'
        
    # Write materials
    tax_rate = float(project.tax_rate) if project.tax_rate else 1.06
    
    for idx, mat in enumerate(materials):
        if idx < len(valid_rows):
            r = valid_rows[idx]
        else:
            r = valid_rows[-1] + (idx - len(valid_rows) + 1)
            
        col_val = lambda col_name, val: ws_project.cell(row=r, column=col_to_num(cols[col_name]), value=val)
        
        col_val("type", mat.type)
        col_val("qty", mat.qty)
        col_val("thickness", mat.thickness)
        if mat.thickness is not None and mat.width is not None:
            col_val("x", "X")
        else:
            col_val("x", None)
        col_val("width", mat.width)
        col_val("length", mat.length)
        col_val("material_type", mat.material_type)
        col_val("lf_pcs", mat.lf_pcs)
        col_val("bf_sf", mat.bf_sf)
        col_val("cost_mbf", mat.cost_mbf)
        col_val("invoice_num", mat.invoice_refs)
        
        # Zero out CO columns
        for c in range(co_start_idx, co_end_idx + 1):
            ws_project.cell(row=r, column=c, value=0.0)
            
        # Write CO adjustments
        mat_co_adjs = [adj for adj in co_adjustments if adj.material_id == mat.id]
        for adj in mat_co_adjs:
            if adj.co_number in co_col_map:
                c = co_col_map[adj.co_number]
                ws_project.cell(row=r, column=c, value=adj.qty_change or 0.0)
                
        # Zero out delivery columns
        for c in range(del_start_idx, del_end_idx + 1):
            ws_project.cell(row=r, column=c, value=0.0)
            
        # Write deliveries
        mat_delivs = [d for d in deliveries if d.material_id == mat.id]
        for d in mat_delivs:
            if d.ship_date:
                dt_norm = datetime(d.ship_date.year, d.ship_date.month, d.ship_date.day)
                if dt_norm in date_col_map:
                    c = date_col_map[dt_norm]
                    curr_val = ws_project.cell(row=r, column=c).value or 0.0
                    ws_project.cell(row=r, column=c, value=curr_val + (d.quantity or 0) * (d.qty_multiplier or 1.0))
                    
        # Inventory
        inv = db.query(Inventory).filter(Inventory.material_id == mat.id).first()
        if inv:
            col_val("inv_bundles", inv.bundles)
            col_val("inv_uom", inv.uom)
            col_val("pcs_bundle", inv.pcs_per_bundle)
            col_val("variance_code", inv.variance_code)
            col_val("reason", inv.reason)
            
        # Formulas
        write_row_formulas_from_scratch(ws_project, r, cols, is_cobia, sorted_dates, del_start_idx, tax_rate)
        
        # Format
        format_data_row(ws_project, r, cols, is_cobia)
        
    # Write intermediate section spacers for Cobia
    write_section_spacers(ws_project, is_cobia)
    
    # Style all headers (row 2)
    for col in range(1, ws_project.max_column + 1):
        cell = ws_project.cell(row=2, column=col)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # VPO sheet
    generate_vpos_sheet(ws_vpos, vpos)
    
    # Autofit column widths
    autofit_columns(ws_project)
    autofit_columns(ws_vpos)
    
    # Ensure export dir exists and save
    os.makedirs(EXPORT_DIR, exist_ok=True)
    wb.save(export_path)
    wb.close()
    
    return export_path
