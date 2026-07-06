"""
Upload router — Phase 2: Intelligent Local Mapping and Heuristics (R2)
"""
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from typing import List, Optional
import os
import shutil
import json
import re

from ..dependencies import get_current_user
from ..models.user import User

from ..database import get_db
from ..config import UPLOAD_DIR, LEGACY_EXCEL
from ..models import Document, Project, COAdjustment, Material, Activity, Delivery, ItemMapping
from ..services.classifier import classify_document
from ..services.pdf_parser import parse_pdf_document

from ..services.matcher import (
    classify_item_category,
    normalize_text,
    parse_dimension_val,
    parse_dimensions_string,
    extract_dimensions_from_text,
    get_dimensions,
    score_match,
    match_material
)

router = APIRouter()

# Ensure upload directory exists
os.makedirs(UPLOAD_DIR, exist_ok=True)


def _log_activity(db: Session, project_id: str, action: str, detail: str):
    """Write an activity log entry."""
    act = Activity(project_id=project_id, action=action, detail=detail)
    db.add(act)


def _save_co_adjustments(db: Session, project_id: str, doc_data, doc_id: str):
    """
    Parse CO line items and persist as COAdjustment records linked to matching Materials.
    Updates existing materials or inserts a new Material record if no match is found.
    """
    from ..services.calculator import compute_material_totals

    project = db.query(Project).filter(Project.id == project_id).first()
    tax_rate = project.tax_rate if project else 0.0

    materials = db.query(Material).filter(Material.project_id == project_id).all()
    for item in doc_data.line_items:
        best = None
        best_score = 0
        for mat in materials:
            score = score_match(
                inv_desc=item.description or "",
                inv_dims=item.dimensions or "",
                inv_code=item.item_code or "",
                mat_type=mat.type or "",
                mat_thickness=mat.thickness,
                mat_width=mat.width,
                mat_length=mat.length,
                mat_desc=mat.material_type or ""
            )
            if score > best_score:
                best_score = score
                best = mat

        inv_category = classify_item_category(item.description or "", item.item_code or "")
        threshold = 10 if inv_category in ("each", "invoice") else 15

        if best and best_score >= threshold:
            best.co_qty = (best.co_qty or 0.0) + item.quantity
            best.po_co_qty = (best.qty or 0.0) + (best.co_qty or 0.0)
            totals = compute_material_totals(best, tax_rate)
            best.lf_pcs = totals.get("lf_pcs", 0.0)
            best.bf_sf = totals.get("bf_sf", 0.0)
            best.total_cost = totals.get("total_cost", 0.0)
            best.total_cost_tax = totals.get("total_cost_tax", 0.0)
        else:
            t, w, l = get_dimensions(item.description or "", item.dimensions or "")
            category = classify_item_category(item.description or "", item.item_code or "")
            best = Material(
                project_id=project_id,
                type=category,
                qty=0.0,
                co_qty=item.quantity,
                po_co_qty=item.quantity,
                thickness=t,
                width=w,
                length=l,
                material_type=item.description,
                cost_mbf=item.unit_price or 0.0
            )
            totals = compute_material_totals(best, tax_rate)
            best.lf_pcs = totals.get("lf_pcs", 0.0)
            best.bf_sf = totals.get("bf_sf", 0.0)
            best.total_cost = totals.get("total_cost", 0.0)
            best.total_cost_tax = totals.get("total_cost_tax", 0.0)
            db.add(best)
            db.flush()
            materials.append(best)

        adj = COAdjustment(
            material_id=best.id if best else None,
            co_number=doc_data.number,
            co_date=str(doc_data.date) if doc_data.date else "",
            qty_change=item.quantity,
            description=item.description,
        )
        db.add(adj)


def _match_line_to_material(item, materials, excel_rows: list, db: Optional[Session] = None, project_id: Optional[str] = None) -> dict:
    """
    Match a parsed line item to the closest material and Excel row reference.
    """
    if not project_id and db and materials:
        project_id = getattr(materials[0], 'project_id', None)

    mapped_mat = None
    if db and project_id and item.description:
        mapping = db.query(ItemMapping).filter(
            ItemMapping.project_id == project_id,
            ItemMapping.invoice_description == item.description
        ).first()
        if mapping:
            for m in materials:
                if m.id == mapping.material_id:
                    mapped_mat = m
                    break
            if not mapped_mat:
                mapped_mat = db.query(Material).filter(Material.id == mapping.material_id).first()

    best = None
    best_score = 0
    if mapped_mat:
        best = mapped_mat
        best_score = 100
    else:
        for mat in materials:
            score = score_match(
                inv_desc=item.description or "",
                inv_dims=item.dimensions or "",
                inv_code=item.item_code or "",
                mat_type=mat.type or "",
                mat_thickness=mat.thickness,
                mat_width=mat.width,
                mat_length=mat.length,
                mat_desc=mat.material_type or ""
            )
            if score > best_score:
                best_score = score
                best = mat

    best_excel = None
    excel_score = 0

    item_desc_norm = normalize_text(item.description or "")
    item_t, item_w, item_l = get_dimensions(item.description or "", item.dimensions or "")

    for er in excel_rows:
        e_score = 0
        e_desc = er.get("description") or ""
        e_desc_norm = normalize_text(e_desc)

        inv_cat = classify_item_category(item.description or "", item.item_code or "")
        er_cat = classify_item_category(e_desc, "")
        if er.get("type"):
            er_cat = classify_item_category(er["type"], "")

        if inv_cat == er_cat:
            e_score += 10
        elif inv_cat in er_cat or er_cat in inv_cat:
            e_score += 5

        er_t = parse_dimension_val(str(er.get("thickness"))) if er.get("thickness") is not None else None
        er_w = parse_dimension_val(str(er.get("width"))) if er.get("width") is not None else None
        er_l = parse_dimension_val(str(er.get("length"))) if er.get("length") is not None else None

        if er_t is None or er_w is None or er_l is None:
            er_t_d, er_w_d, er_l_d = get_dimensions(e_desc, "")
            if er_t is None: er_t = er_t_d
            if er_w is None: er_w = er_w_d
            if er_l is None: er_l = er_l_d

        if item_t is not None and er_t is not None:
            if abs(item_t - er_t) < 0.01:
                e_score += 5
        if item_w is not None and er_w is not None:
            if abs(item_w - er_w) < 0.01:
                e_score += 5
        if item_l is not None and er_l is not None:
            if abs(item_l - er_l) < 0.01:
                e_score += 5

        keywords = ["TREATED", "SOUTHERN YELLOW PINE", "LVL", "OSB", "PLYWOOD", "ZIP"]
        for kw in keywords:
            if kw in item_desc_norm and kw in e_desc_norm:
                e_score += 3

        i_words = set(item_desc_norm.split())
        e_words = set(e_desc_norm.split())
        filler = {"X", "AND", "THE", "OF", "FOR", "IN", "TO", "WITH", "FT", "INCH", "PCS", "PC"}
        i_words = i_words - filler
        e_words = e_words - filler
        common = i_words & e_words
        common = {w for w in common if len(w) > 2}
        e_score += len(common) * 2

        if e_score > excel_score:
            excel_score = e_score
            best_excel = er

    inv_category = classify_item_category(item.description or "", item.item_code or "")
    # For lumber, require a length match (score >= 18) when no length is specified in the
    # invoice line. This prevents an unspecific "2x4 SYP#2" invoice line from dumping all
    # its delivered quantity onto the first 2x4 SYP#2 row regardless of length.
    item_t, item_w, item_l = get_dimensions(item.description or "", item.dimensions or "")
    lumber_no_length = (inv_category == "lumber" and (item_l is None or item_l == 0))
    threshold = 10 if inv_category in ("each", "invoice") else (18 if lumber_no_length else 15)
    if mapped_mat:
        threshold = 0

    matched = (best is not None) and (best_score >= threshold)

    qty_multiplier = 1.0
    if matched and best:
        if best.length and item_l and best.length != item_l:
            qty_multiplier = item_l / best.length

    return {
        "matched": matched,
        "score": max(best_score, excel_score),
        "material_id": best.id if (best and best_score >= threshold) else None,
        "material_type": best.material_type if (best and best_score >= threshold) else None,
        "material_dimensions": f"{best.thickness or ''}x{best.width or ''}x{best.length or ''}" if (best and best_score >= threshold) else None,
        "excel_row": best_excel,
        "qty_multiplier": qty_multiplier,
    }


_excel_row_refs_cache = {}


def _load_excel_row_refs(project_id: str, db: Optional[Session] = None) -> list:
    """
    Load lightweight row references from Client_Requirments_Doc.xlsx.
    Returns list of {row, sheet, type, description, ...} dicts.
    """
    project_name = ""
    if db is not None:
        try:
            project = db.query(Project).filter(Project.id == project_id).first()
            if project:
                project_name = project.name
        except Exception:
            pass
    if not project_name:
        project_name = project_id
    if project_name in _excel_row_refs_cache:
        return _excel_row_refs_cache[project_name]
    if not os.path.exists(LEGACY_EXCEL):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(LEGACY_EXCEL, data_only=True, read_only=True)
        name_upper = project_name.upper()
        sheet_name = None

        for sn in wb.sheetnames:
            su = sn.upper()
            if "COBIA" in name_upper and "COBIA" in su:
                sheet_name = sn
                break
            if "WILLOW" in name_upper and "WILLOW" in su:
                sheet_name = sn
                break

        # If project name is unrecognized, search sheets case-insensitively
        if not sheet_name:
            for sn in wb.sheetnames:
                su = sn.upper()
                if name_upper in su or su in name_upper:
                    sheet_name = sn
                    break
            # Fallback to the first sheet (ignoring "VPO's")
            if not sheet_name:
                valid_sheets = [sn for sn in wb.sheetnames if "VPO" not in sn.upper()]
                if valid_sheets:
                    sheet_name = valid_sheets[0]
                elif wb.sheetnames:
                    sheet_name = wb.sheetnames[0]

        if not sheet_name:
            wb.close()
            return []

        ws = wb[sheet_name]

        # Scan header row (row 2, falling back to row 1)
        header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), [])
        header_row_num = 2
        key_words = ["type", "material", "description", "thickness", "width", "length", "qty", "cost"]
        row2_has_keywords = any(any(kw in str(val).lower() for kw in key_words) for val in header_row if val)
        if not row2_has_keywords:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            header_row_num = 1

        # Default indices
        type_col_idx = 1
        mat_col_idx = 23
        thick_col_idx = None
        width_col_idx = None
        length_col_idx = None
        qty_col_idx = None
        cost_col_idx = None

        for idx, val in enumerate(header_row):
            if not val:
                continue
            val_str = str(val).strip().lower()

            if "material type" in val_str or "description" in val_str or val_str == "desc" or "material_type" in val_str:
                mat_col_idx = idx + 1
            elif val_str == "type" or "division" in val_str or "category" in val_str:
                type_col_idx = idx + 1
            elif "thickness" in val_str or val_str == "thick" or val_str == "t":
                thick_col_idx = idx + 1
            elif "width" in val_str or val_str == "w":
                width_col_idx = idx + 1
            elif "length" in val_str or val_str == "l" or val_str == "len":
                length_col_idx = idx + 1
            elif "qty" in val_str or "quantity" in val_str or val_str == "q":
                qty_col_idx = idx + 1
            elif "cost" in val_str or "price" in val_str or "rate" in val_str:
                cost_col_idx = idx + 1

        rows = []
        for row_num in range(header_row_num + 1, 200):
            type_val = ws.cell(row=row_num, column=type_col_idx).value
            mat_val = ws.cell(row=row_num, column=mat_col_idx).value

            thick_val = ws.cell(row=row_num, column=thick_col_idx).value if thick_col_idx else None
            width_val = ws.cell(row=row_num, column=width_col_idx).value if width_col_idx else None
            length_val = ws.cell(row=row_num, column=length_col_idx).value if length_col_idx else None
            qty_val = ws.cell(row=row_num, column=qty_col_idx).value if qty_col_idx else None
            cost_val = ws.cell(row=row_num, column=cost_col_idx).value if cost_col_idx else None

            if type_val and str(type_val).strip():
                rows.append({
                    "row": row_num,
                    "sheet": sheet_name,
                    "type": str(type_val).strip(),
                    "description": str(mat_val).strip() if mat_val else str(type_val).strip(),
                    "thickness": thick_val,
                    "width": width_val,
                    "length": length_val,
                    "qty": qty_val,
                    "cost": cost_val,
                })
        wb.close()
        _excel_row_refs_cache[project_name] = rows
        return rows
    except Exception as e:
        print(f"Excel row ref load error: {e}")
        return []


@router.post("/")
async def upload_files(
    files: List[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Step 1: Upload PDFs. Returns classification result for each file.
    Files are saved to UPLOAD_DIR for confirmation in step 2.
    """
    results = []
    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            results.append({"filename": file.filename, "error": "Only PDFs accepted"})
            continue

        safe_name = os.path.basename(file.filename)
        filepath = os.path.join(UPLOAD_DIR, safe_name)
        with open(filepath, "wb") as f:
            shutil.copyfileobj(file.file, f)

        import pdfplumber
        try:
            with pdfplumber.open(filepath) as pdf:
                text = "\n".join(
                    (p.extract_text(layout=True) or "") for p in pdf.pages
                )
        except Exception as e:
            text = ""

        classification = classify_document(text)
        results.append({
            "filename": safe_name,
            "classification": classification,
        })

    return {"message": f"Uploaded {len(results)} files.", "results": results}


@router.post("/preview")
async def preview_upload(
    file: UploadFile = File(...),
    doc_type: str = Form(...),
    project_id: str = Form(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Step 1.5: Preview what will change if this document is confirmed.
    Returns a rich diff mapped to Client_Requirments_Doc.xlsx rows.
    """
    safe_name = os.path.basename(file.filename)
    filepath = os.path.join(UPLOAD_DIR, safe_name)
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)

    if project_id.startswith("demo-"):
        project = Project(id="demo-0", name="Demo Project", organization_id=current_user.organization_id)
        materials = []
    else:
        project = db.query(Project).filter(
            Project.id == project_id,
            Project.organization_id == current_user.organization_id
        ).first()
        materials = db.query(Material).filter(Material.project_id == project_id).all()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        doc_data = parse_pdf_document(filepath, doc_type)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse PDF: {str(e)}")

    excel_rows = _load_excel_row_refs(project_id, db)

    preview_items = []
    for item in doc_data.line_items:
        match = _match_line_to_material(item, materials, excel_rows, db=db, project_id=project_id)

        if doc_type == "PO":
            if match["matched"]:
                change_type = "PO_UPDATE"
                change_label = "Update existing requirement"
                change_color = "blue"
            else:
                change_type = "ADD"
                change_label = "Will be added to project"
                change_color = "green"
        elif doc_type == "INV":
            if match["matched"]:
                change_type = "INVOICE"
                change_label = "Billing existing material"
                change_color = "blue"
            else:
                change_type = "NEW_CHARGE"
                change_label = "⚠ New charge — no matching PO line"
                change_color = "yellow"
        elif doc_type == "CO":
            if item.quantity > 0:
                change_type = "CO_ADD"
                change_label = f"+{int(item.quantity)} qty adjustment"
                change_color = "green"
            elif item.quantity < 0:
                change_type = "CO_REMOVE"
                change_label = f"{int(item.quantity)} qty reduction"
                change_color = "red"
            else:
                change_type = "CO_ADJUST"
                change_label = "Change order adjustment"
                change_color = "yellow"
        else:
            change_type = "UNKNOWN"
            change_label = "Unknown change"
            change_color = "gray"

        excel_ref = None
        if match.get("excel_row"):
            er = match["excel_row"]
            excel_ref = f"Row {er['row']} · {er['sheet']} · {er['description']}"

        preview_items.append({
            "line_item": {
                "quantity": item.quantity,
                "uom": item.uom,
                "item_code": item.item_code,
                "description": item.description,
                "footage": item.footage,
                "unit_price": item.unit_price,
                "amount": item.amount,
                "dimensions": item.dimensions,
            },
            "change_type": change_type,
            "change_label": change_label,
            "change_color": change_color,
            "matched_material_id": match["material_id"],
            "matched_material_type": match["material_type"],
            "matched_dimensions": match["material_dimensions"],
            "excel_row_ref": excel_ref,
            "match_score": match["score"],
            "match_score_pct": min(100, int((match["score"] / 20.0) * 100)) if match["score"] > 0 else 0,
        })

    duplicate_warning = None
    if doc_type == "INV" and doc_data.number:
        existing = db.query(Document).filter(
            Document.project_id == project_id,
            Document.doc_number == doc_data.number,
        ).first()
        if existing:
            duplicate_warning = f"Invoice #{doc_data.number} has already been processed for this project."

    summary_map = {
        "PO": f"{len(preview_items)} materials will be added to the project",
        "INV": f"{len(preview_items)} line items will be recorded as invoiced",
        "CO": f"{len(preview_items)} quantity adjustments will be applied",
    }

    return {
        "filename": safe_name,
        "doc_type": doc_type,
        "doc_number": doc_data.number,
        "doc_date": str(doc_data.date) if doc_data.date else None,
        "project_name": doc_data.project_name,
        "subtotal": doc_data.subtotal,
        "tax": doc_data.tax,
        "total_amount": doc_data.total_amount,
        "line_items_count": len(preview_items),
        "preview_items": preview_items,
        "duplicate_warning": duplicate_warning,
        "excel_available": os.path.exists(LEGACY_EXCEL),
        "summary": summary_map.get(doc_type, f"{len(preview_items)} changes will be applied"),
    }


@router.post("/confirm")
async def confirm_upload(
    file: UploadFile = File(...),
    doc_type: str = Form(...),
    project_id: str = Form(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Step 2: Confirm a previously uploaded file for processing.
    """
    safe_name = os.path.basename(file.filename)
    filepath = os.path.join(UPLOAD_DIR, safe_name)
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)

    if project_id.startswith("demo-"):
        try:
            doc_data = parse_pdf_document(filepath, doc_type)
        except Exception:
            doc_data = None

        return {
            "message": "Demo Document processed successfully (No DB changes)",
            "document_id": 999,
            "doc_number": doc_data.number if doc_data else "DEMO-123",
            "line_items_parsed": len(doc_data.line_items) if doc_data else 0,
        }

    project = db.query(Project).filter(Project.id == project_id, Project.organization_id == current_user.organization_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        doc_data = parse_pdf_document(filepath, doc_type)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse PDF: {str(e)}")

    if doc_type == "INV" and doc_data.number:
        existing = db.query(Document).filter(
            Document.project_id == project_id,
            Document.doc_number == doc_data.number,
        ).first()
        if existing:
            raise HTTPException(
                status_code=409,
                detail=f"Invoice #{doc_data.number} has already been processed for this project.",
            )

    try:
        doc = Document(
            project_id=project_id,
            doc_type=doc_type,
            filename=safe_name,
            doc_number=doc_data.number,
            parsed_data_json=doc_data.model_dump_json(),
        )
        db.add(doc)
        db.flush()

        if doc_type == "CO":
            _save_co_adjustments(db, project_id, doc_data, doc.id)

        elif doc_type == "PO":
            from ..services.calculator import compute_material_totals
            materials_list = db.query(Material).filter(Material.project_id == project_id).all()
            excel_rows = _load_excel_row_refs(project_id, db)
            for item in doc_data.line_items:
                match_res = _match_line_to_material(item, materials_list, excel_rows, db=db, project_id=project_id)
                if match_res["matched"] and match_res["material_id"]:
                    mat = db.query(Material).filter(Material.id == match_res["material_id"]).first()
                    if mat:
                        mat.qty = (mat.qty or 0) + item.quantity
                        mat.po_co_qty = (mat.qty or 0) + (mat.co_qty or 0)
                        totals = compute_material_totals(mat, project.tax_rate)
                        mat.lf_pcs = totals.get("lf_pcs", 0.0)
                        mat.bf_sf = totals.get("bf_sf", 0.0)
                        mat.total_cost = totals.get("total_cost", 0.0)
                        mat.total_cost_tax = totals.get("total_cost_tax", 0.0)
                        continue

                # Use item_code directly if it's a known PO category — avoids misclassification
                KNOWN_PO_TYPES = {"LVL": "lvl", "LUMBER": "lumber", "PANELS": "panels", "EACH": "each"}
                if item.item_code.upper() in KNOWN_PO_TYPES:
                    category = KNOWN_PO_TYPES[item.item_code.upper()]
                else:
                    category = classify_item_category(item.description or "", item.item_code or "")
                mat_desc = item.description
                if match_res.get("matched") and match_res.get("excel_row"):
                    excel_row = match_res["excel_row"]
                    if excel_row.get("type"):
                        category = excel_row["type"]

                t, w, l = get_dimensions(item.description or "", item.dimensions or "")
                mat = Material(
                    project_id=project_id,
                    type=category,
                    qty=item.quantity,
                    co_qty=0.0,
                    po_co_qty=item.quantity,
                    thickness=t,
                    width=w,
                    length=l,
                    material_type=mat_desc,
                    cost_mbf=item.unit_price,
                )
                totals = compute_material_totals(mat, project.tax_rate)
                mat.lf_pcs = totals.get("lf_pcs", 0.0)
                mat.bf_sf = totals.get("bf_sf", 0.0)
                mat.total_cost = totals.get("total_cost", 0.0)
                mat.total_cost_tax = totals.get("total_cost_tax", 0.0)
                db.add(mat)

        elif doc_type == "INV":
            materials = db.query(Material).filter(Material.project_id == project_id).all()
            excel_rows = _load_excel_row_refs(project_id, db)
            for item in doc_data.line_items:
                match_res = _match_line_to_material(item, materials, excel_rows, db=db, project_id=project_id)
                if match_res["matched"] and match_res["material_id"]:
                    matched_material_id = match_res["material_id"]
                    deliv = Delivery(
                        material_id=matched_material_id,
                        document_id=doc.id,
                        invoice_number=doc_data.number,
                        ship_date=doc_data.date,
                        quantity=item.quantity,
                        qty_multiplier=match_res.get("qty_multiplier", 1.0),
                        uom=item.uom,
                    )
                    db.add(deliv)

                    # Update Material's invoice_refs by appending invoice number
                    mat = db.query(Material).filter(Material.id == matched_material_id).first()
                    if mat and doc_data.number:
                        import re as re_inv
                        current_refs = re_inv.split(r'[,\n]+', mat.invoice_refs or "")
                        existing_refs = [r.strip() for r in current_refs if r.strip()]
                        if doc_data.number not in existing_refs:
                            existing_refs.append(doc_data.number)
                            mat.invoice_refs = ", ".join(existing_refs)

        _log_activity(
            db, project_id,
            action=f"Document Processed: {doc_type}",
            detail=f"{safe_name} | Doc# {doc_data.number} | {len(doc_data.line_items)} line items",
        )
        db.commit()
        db.refresh(doc)

    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Database transaction failed: {str(e)}")

    return {
        "message": "Document processed and confirmed successfully",
        "document_id": doc.id,
        "doc_number": doc_data.number,
        "line_items_parsed": len(doc_data.line_items),
    }
